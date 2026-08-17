"""Export sample data to Excel — same format as Sample_DB_E16.xlsx, v2 schema."""
import os, json
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2
from psycopg2.extras import RealDictCursor

from dotenv import load_dotenv
load_dotenv()

DB_URL = os.getenv("DATABASE_URL", "postgresql://contractlens:contractlens@localhost:5433/contractlens")
COMMENT_COL = 21  # fixed comment column (col U)

# ── column meanings (Vietnamese) ──────────────────────────────────────────
COL_MEANINGS = {
    "users": {
        "id": "UUID PK", "email": "Đăng nhập", "password_hash": "Bcrypt hash",
        "created_at": "Thời điểm đăng ký",
    },
    "uploaded_contracts": {
        "contract_id": "PK TEXT", "user_id": "FK → users.id",
        "filename": "Tên file gốc", "file_type": "docx|pdf|…",
        "storage_key": "Key lưu trữ (MinIO)", "full_text": "Nguyên văn SoT",
        "status": "Trạng thái pipeline", "message": "Thông báo UI",
        "chunk_count": "Số chunk đã embed", "analysis": "JSONB extract",
        "risks": "JSONB risk items", "created_at": "Tạo", "updated_at": "Cập nhật",
    },
    "contract_chunks": {
        "id": "BIGSERIAL", "contract_id": "FK → uploaded_contracts",
        "chunk_index": "Thứ tự cắt", "clause_number": "Số Điều",
        "content": "Nguyên văn chunk", "embedding": "vector(1024) HNSW",
    },
    "legal_documents": {
        "doc_id": "PK TEXT (vbpl id)", "doc_num": "Số hiệu hiển thị",
        "title": "Tiêu đề đầy đủ", "doc_type": "Nghị định|Luật|Thông tư|…",
        "majors": "TEXT[] ngành", "fields": "TEXT[] lĩnh vực",
        "issue_date": "DATE ban hành", "eff_from": "Ngày HL khai báo",
        "eff_to": "Hết HL (nullable)", "eff_flag": "Nhãn hiệu lực (9 giá trị)",
        "status_flag": "Lọc RAG (0=? 1=còn 2=hết 3=chưa 4=hết1P 5=còn1P)",
        "agency": "Cơ quan BH", "signer_name": "Người ký",
        "signer_title": "Chức danh người ký", "source_url": "URL nguồn VBPL",
        "full_text": "TEXT tùy chọn", "path": "LTREE root = sanitized doc_id",
        "crawled_at": "Thời điểm crawl", "updated_at": "Cập nhật",
    },
    "legal_parts": {
        "id": "BIGSERIAL", "doc_id": "FK → legal_documents (chỉ bảng hierarchy có doc_id)",
        "title": "Tiêu đề; thiếu cấp → 'Không có'",
        "content": "Nội dung; thiếu → 'Không có'",
        "path": "LTREE (thật P* hoặc scaffold ._P)", "parent_path": "LTREE = doc path",
        "eff_from": "Ngày HL (inherit doc nếu NULL)", "eff_to": "Hết HL",
        "eff_flag": "Nhãn HL (đồng bộ status_flag)", "status_flag": "0..5 như legal_documents",
        "created_at": "Tạo",
    },
    "legal_chapters": {
        "id": "BIGSERIAL",
        "part_id": "FK → legal_parts.id NOT NULL",
        "title": "Tiêu đề Chương; thiếu → 'Không có'",
        "content": "Nội dung; thiếu → 'Không có'",
        "path": "LTREE UNIQUE (C* hoặc ._C)", "parent_path": "LTREE → part",
        "eff_from": "Ngày HL", "eff_to": "Hết HL",
        "eff_flag": "Nhãn HL", "status_flag": "0..5; cha=2 cascade xuống",
        "created_at": "Tạo",
    },
    "legal_sections": {
        "id": "BIGSERIAL",
        "chapter_id": "FK → legal_chapters.id NOT NULL",
        "title": "Tiêu đề Mục; thiếu → 'Không có'",
        "content": "Nội dung; thiếu → 'Không có'",
        "path": "LTREE UNIQUE (M* hoặc ._M)", "parent_path": "LTREE → chapter",
        "eff_from": "Ngày HL", "eff_to": "Hết HL",
        "eff_flag": "Nhãn HL", "status_flag": "0..5",
        "created_at": "Tạo",
    },
    "legal_sub_sections": {
        "id": "BIGSERIAL",
        "section_id": "FK → legal_sections.id NOT NULL",
        "title": "Tiêu đề Tiểu mục; thiếu → 'Không có'",
        "content": "Nội dung; thiếu → 'Không có'",
        "path": "LTREE UNIQUE (TM* hoặc ._TM)", "parent_path": "LTREE → section",
        "eff_from": "Ngày HL", "eff_to": "Hết HL",
        "eff_flag": "Nhãn HL", "status_flag": "0..5",
        "created_at": "Tạo",
    },
    "legal_articles": {
        "id": "BIGSERIAL",
        "sub_section_id": "FK → legal_sub_sections.id NOT NULL (luôn có chuỗi FK)",
        "title": "Tiêu đề Điều NOT NULL",
        "content": "Nội dung; thiếu → 'Không có'",
        "path": "LTREE UNIQUE (D*)", "parent_path": "LTREE → sub_section",
        "eff_from": "Ngày HL (ex eff_date)", "eff_to": "Hết HL",
        "eff_flag": "Nhãn HL", "status_flag": "0..5; hết HL toàn bộ cascade Khoản/Điểm",
        "created_at": "Tạo",
    },
    "legal_clauses": {
        "id": "BIGSERIAL",
        "article_id": "FK → legal_articles.id NOT NULL",
        "title": "Tiêu đề Khoản NOT NULL",
        "content": "Nội dung; thiếu → 'Không có'",
        "path": "LTREE UNIQUE (K*)", "parent_path": "LTREE → article",
        "eff_from": "Ngày HL", "eff_to": "Hết HL",
        "eff_flag": "Nhãn HL", "status_flag": "0..5; partial không đụng sibling",
        "created_at": "Tạo",
    },
    "legal_points": {
        "id": "BIGSERIAL",
        "clause_id": "FK → legal_clauses.id NOT NULL",
        "symbol": "Ký hiệu ltree leaf (a, b, c, dd, …)",
        "title": "Tiêu đề; thiếu → 'Không có'",
        "content": "Nội dung; thiếu → 'Không có'",
        "path": "LTREE UNIQUE",
        "parent_path": "LTREE → clause",
        "eff_from": "Ngày HL", "eff_to": "Hết HL",
        "eff_flag": "Nhãn HL", "status_flag": "0..5",
        "created_at": "Tạo",
    },
    "legal_embeddings": {
        "id": "BIGSERIAL",
        "doc_id": "FK → legal_documents", "chunk_type": "body|preamble|…",
        "chunk_text": "Text embed + FTS + RAG", "embedding": "vector(1024) BGE-M3",
        "is_effective": "Cache RAG (hierarchy status IN 1,5); SoT = hierarchy",
        "tsv": "TSVECTOR GENERATED",
        "path": "LTREE UNIQUE leaf key (sanitize(doc_id).structural)",
        "root_path": "LTREE tới Điều (D*) gần nhất",
        "created_at": "Tạo",
    },
    "legal_document_relations": {
        "id": "BIGSERIAL", "from_doc_id": "FK nguồn → legal_documents",
        "to_doc_id": "FK đích → legal_documents",
        "relation_type": "Mã quan hệ luoc_do (can_cu|sua_doi|thay_the|…)",
        "created_at": "Tạo",
    },
    "legal_path_relations": {
        "id": "BIGSERIAL", "source_path": "LTREE nút trích dẫn (Điều/Khoản/Điểm)",
        "target_path": "LTREE nút được dẫn chiếu",
        "ref_type": "dan_chieu|can_cu|sua_doi|…",
        "created_at": "Tạo",
    },
}

# ── Table descriptions (col U) and per-row comments ────────────────────────
TABLE_NOTES = {
    "users": "Bảng tài khoản local (JWT). PK: id (UUID). email UNIQUE. password_hash = bcrypt. FK từ uploaded_contracts.user_id.",
    "uploaded_contracts": "Hợp đồng user upload. PK: contract_id. storage_key = đường dẫn file local/MinIO. analysis/risks = JSONB cache.",
    "contract_chunks": "Chunk hợp đồng theo Điều. UNIQUE(contract_id, chunk_index). clause_number = số Điều. embedding = vector(1024) BGE-M3 HNSW.",
    "legal_documents": "Metadata thuoc_tinh. PK: doc_id. path = LTREE root. Chỉ legal_parts (+ embeddings) FK trực tiếp doc_id.",
    "legal_parts": "Cấp 1: Phần. Mọi VB ≥1 row (thật hoặc title/content='Không có'). Duy nhất hierarchy có doc_id. Có eff_*/status_flag.",
    "legal_chapters": "Cấp 2: Chương. part_id NOT NULL; thiếu Phần → scaffold ._P. status=2 cascade xuống Mục.",
    "legal_sections": "Cấp 3: Mục. chapter_id NOT NULL; thiếu → scaffold ._C + 'Không có'.",
    "legal_sub_sections": "Cấp 4: Tiểu mục. section_id NOT NULL; thiếu → scaffold ._M/._TM + 'Không có'.",
    "legal_articles": "Cấp 5: Điều. Luôn sub_section_id NOT NULL. Hết HL toàn bộ → cascade Khoản/Điểm.",
    "legal_clauses": "Cấp 6: Khoản. article_id NOT NULL. Partial repeal không đụng sibling.",
    "legal_points": "Cấp 7: Điểm. symbol = a/b/dd. clause_id NOT NULL. title/content default 'Không có'.",
    "legal_embeddings": "RAG SoT text+embedding+FTS. is_effective = cache từ hierarchy status IN (1,5).",
    "legal_document_relations": "Cạnh doc↔doc từ luoc_do (bãi bỏ/thay thế/căn cứ…). UNIQUE(from_doc_id, to_doc_id, relation_type).",
    "legal_path_relations": "Cạnh dẫn chiếu path↔path (dan_chieu Điều/Khoản/Điểm). Song song legal_document_relations.",
}

# Per-row comments (list of strings, one per sample row)
ROW_COMMENTS = {
    "users": [
        "User mẫu A — có 2 hợp đồng bên dưới",
        "User mẫu B — có hợp đồng pending",
        "User mẫu C — chưa upload hợp đồng nào",
    ],
    "uploaded_contracts": [
        "Đã analyze hoàn tất — 3 chunk embedding",
        "Đã analyze — 2 chunk (thuê VP)",
        "Mới upload — đang pending, chưa embed",
    ],
    "contract_chunks": [
        "Điều 2 HĐLĐ — map judge theo clause_number",
        "Điều 4 thử việc — đối chiếu BLLĐ Điều 36",
        "Điều 7 rủi ro — đơn phương chấm dứt",
        "HĐ thuê — Điều 1 đối tượng thuê",
        "HĐ thuê — Điều 3 giá thuê",
    ],
    "legal_documents": [
        "NĐ 168 — Luật GTĐB, hiệu lực 2025 (sample chính)",
        "BLLĐ 2019 — hiệu lực 2021",
        "NĐ 100 — hết HL một phần (bị NĐ 168 sửa đổi)",
        "VB chưa có hiệu lực — eff_from tương lai",
    ],
    "legal_parts": [
        "Part thật (P*) hoặc scaffold ._P title='Không có'",
        "Part thật khác cùng doc",
    ],
    "legal_chapters": [
        "Chương thật (C1) — part_id NOT NULL",
        "Scaffold ._C nếu VB không có Chương",
    ],
    "legal_sections": [
        "Mục thật (M1) — chapter_id NOT NULL",
        "Scaffold ._M nếu VB không có Mục",
    ],
    "legal_sub_sections": [
        "Tiểu mục TM1 dưới Mục",
        "Scaffold ._TM nếu VB không có Tiểu mục",
    ],
    "legal_articles": [
        "Điều thật (D1) — luôn dưới sub_section",
        "Điều thật (D2)",
    ],
    "legal_clauses": [
        "Khoản thật (K1) — article_id NOT NULL",
        "Khoản thật (K2)",
    ],
    "legal_points": [
        "Điểm a — symbol + title/content",
        "Điểm b — sibling cùng Khoản",
        "Điểm dd (chữ đ) — ký hiệu đặc biệt",
    ],
    "legal_embeddings": [
        "Chunk body — Điều+Khoản+Điểm gộp chung",
        "Chunk preamble — căn cứ ban hành",
        "Chunk effectivity — hiệu lực thi hành",
    ],
    "legal_document_relations": [
        "168 BASED_ON Luật XLHC",
        "168 AMENDS NĐ 100 (sửa đổi bổ sung)",
        "BLLĐ CITES Luật XLHC (mẫu)",
    ],
    "legal_path_relations": [
        "Điểm b gắn ngữ cảnh với điểm a (sibling)",
        "Khoản 2 tham chiếu Khoản 1 cùng Điều",
        "Hiệu lực áp dụng toàn NĐ kể cả Điều 1 (kèm ngày HL)",
        "Tham chiếu chéo: NĐ 168 → BLLĐ 2019",
    ],
}

# ── helpers ────────────────────────────────────────────────────────────────
YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True, size=11)
NORMAL = Font(size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


def _fmt(v):
    """Format cell value for Excel."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return str(v)
    if isinstance(v, (list, dict)):
        s = json.dumps(v, ensure_ascii=False)
        return s[:300]
    if isinstance(v, float):
        return round(v, 4)
    if isinstance(v, bytes):
        return f"<binary {len(v)}b>"
    return str(v)[:500]


def fake_contract_chunks():
    """Generate fake sample rows for contract_chunks (DB has 0 rows)."""
    return [
        {"id": 1, "contract_id": "ctr-hdl-2026-001", "chunk_index": 1, "clause_number": "2",
         "content": "Điều 2. Tiền lương…\nMức lương: 12.000.000 VNĐ/tháng, trả ngày 05 hàng tháng.",
         "embedding": "<vector(1024)>"},
        {"id": 2, "contract_id": "ctr-hdl-2026-001", "chunk_index": 2, "clause_number": "4",
         "content": "Điều 4. Thời gian thử việc\nThử việc 02 tháng, hưởng 85% lương chính thức.",
         "embedding": "<vector(1024)>"},
        {"id": 3, "contract_id": "ctr-hdl-2026-001", "chunk_index": 3, "clause_number": "7",
         "content": "Điều 7. Đơn phương chấm dứt\nVi phạm bồi thường nửa tháng lương.",
         "embedding": "<vector(1024)>"},
        {"id": 4, "contract_id": "ctr-thue-2026-002", "chunk_index": 1, "clause_number": "1",
         "content": "Điều 1. Đối tượng thuê\nMặt bằng tầng 5, 80m2.",
         "embedding": "<vector(1024)>"},
        {"id": 5, "contract_id": "ctr-thue-2026-002", "chunk_index": 2, "clause_number": "3",
         "content": "Điều 3. Giá thuê\n25.000.000 VNĐ/tháng.",
         "embedding": "<vector(1024)>"},
    ]


def connect():
    return psycopg2.connect(DB_URL, cursor_factory=RealDictCursor)


def table_columns(cur, table_name):
    cur.execute("""
        SELECT column_name, ordinal_position
        FROM information_schema.columns
        WHERE table_schema='public' AND table_name=%s
        ORDER BY ordinal_position
    """, (table_name,))
    return [r["column_name"] for r in cur.fetchall()]


def sample_rows(cur, table_name, where="", limit=4, order=""):
    o = order if order else f"ORDER BY RANDOM()"
    q = f"SELECT * FROM {table_name} {where} {o} LIMIT {limit}"
    try:
        cur.execute(q)
        return cur.fetchall()
    except Exception as e:
        print(f"  WARN {table_name}: {e}")
        return []


# ── build workbook ─────────────────────────────────────────────────────────
def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample du lieu v2"
    conn = connect()
    cur = conn.cursor()

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COMMENT_COL - 2)
    ws.cell(1, 1, "Sample dữ liệu Database v2 (Postgres + ltree hierarchy)").font = Font(bold=True, size=14)
    ws.cell(1, COMMENT_COL, "CỘT CHÚ THÍCH (cố định bên phải)").font = BOLD
    ws.cell(1, COMMENT_COL).fill = YELLOW

    # Row 2: subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COMMENT_COL - 2)
    ws.cell(2, 1, "Các bảng xếp liên tục từ trên xuống · Cột bên phải (vàng) = chú thích · SoT: hierarchy 7 bảng + legal_embeddings.").font = Font(size=9, italic=True)

    row = 4  # start from row 4

    tables = [
        "users", "uploaded_contracts", "contract_chunks",
        "legal_documents", "legal_parts", "legal_chapters", "legal_sections",
        "legal_sub_sections", "legal_articles", "legal_clauses", "legal_points",
        "legal_embeddings", "legal_document_relations", "legal_path_relations",
    ]

    for idx, tbl in enumerate(tables, 1):
        cols = table_columns(cur, tbl)
        meanings = COL_MEANINGS.get(tbl, {})
        notes = TABLE_NOTES.get(tbl, "")
        row_comments = ROW_COMMENTS.get(tbl, [])

        # Fetch sample rows with custom filtering
        where = ""
        order = "ORDER BY RANDOM()"
        limit = 4
        if tbl == "legal_documents":
            rows = sample_rows(cur, tbl, "WHERE status_flag=1", limit=2, order="ORDER BY RANDOM()")
            rows += sample_rows(cur, tbl, "WHERE status_flag!=1 AND status_flag!=0", limit=2, order="ORDER BY RANDOM()")
        elif tbl == "legal_path_relations":
            rows = []  # empty — will insert fake
        elif tbl == "contract_chunks":
            rows = sample_rows(cur, tbl, limit=limit)
            if not rows:
                rows = fake_contract_chunks()
        elif tbl == "legal_embeddings":
            rows = sample_rows(cur, tbl, "WHERE embedding IS NOT NULL", limit=3, order="ORDER BY RANDOM()")
        else:
            rows = sample_rows(cur, tbl, limit=limit)

        # ── row: table header ──
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COMMENT_COL - 2)
        ws.cell(row, 1, f"{idx}. {tbl}").font = BOLD
        ws.cell(row, COMMENT_COL, notes).font = Font(size=9)
        ws.cell(row, COMMENT_COL).fill = YELLOW
        ws.cell(row, COMMENT_COL).alignment = WRAP
        for c in range(1, COMMENT_COL + 1):
            ws.cell(row, c).border = BORDER
        row += 1

        # ── row: column names ──
        for ci, col in enumerate(cols, 1):
            ws.cell(row, ci, col).font = Font(bold=True, size=10)
            ws.cell(row, ci).border = BORDER
        ws.cell(row, COMMENT_COL, "Chú thích").font = BOLD
        ws.cell(row, COMMENT_COL).fill = YELLOW
        ws.cell(row, COMMENT_COL).border = BORDER
        row += 1

        # ── row: column meanings ──
        for ci, col in enumerate(cols, 1):
            ws.cell(row, ci, meanings.get(col, "")).font = Font(size=9, italic=True, color="666666")
            ws.cell(row, ci).border = BORDER
        ws.cell(row, COMMENT_COL, "← ý nghĩa từng cột (không phải dữ liệu DB)").font = Font(size=9, italic=True, color="999999")
        ws.cell(row, COMMENT_COL).fill = YELLOW
        ws.cell(row, COMMENT_COL).border = BORDER
        row += 1

        # ── data rows ──
        if not rows and tbl == "legal_path_relations":
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COMMENT_COL - 2)
            ws.cell(row, 1, "Chưa có dữ liệu thật — xem sample fake bên dưới.").font = Font(
                size=9, italic=True, color="999999"
            )
            for c in range(1, COMMENT_COL + 1):
                ws.cell(row, c).border = BORDER
            row += 1

        for ri, r in enumerate(rows):
            for ci, col in enumerate(cols, 1):
                val = r.get(col)
                if col in ("password_hash", "embedding"):
                    val = "***REDACTED***"
                elif col == "tsv" and val:
                    val = "<tsvector>"
                elif col in ("chunk_content", "chunk_text") and val:
                    val = str(val)[:120] + ("…" if len(str(val)) > 120 else "")
                elif col in ("full_text", "content") and val:
                    val = str(val)[:200] + ("…" if len(str(val)) > 200 else "")
                ws.cell(row, ci, _fmt(val)).font = NORMAL
                ws.cell(row, ci).border = BORDER
                ws.cell(row, ci).alignment = WRAP
            comment = row_comments[ri] if ri < len(row_comments) else ""
            ws.cell(row, COMMENT_COL, comment).font = Font(size=9)
            ws.cell(row, COMMENT_COL).fill = YELLOW
            ws.cell(row, COMMENT_COL).alignment = WRAP
            ws.cell(row, COMMENT_COL).border = BORDER
            row += 1

        # Insert fake legal_path_relations after the empty row
        if tbl == "legal_path_relations":
            fake_refs = [
                {"id": 1, "source_path": "168_2024_nd_cp.C2.D6.K3.b", "target_path": "168_2024_nd_cp.C2.D6.K3.a", "ref_type": "dan_chieu", "created_at": "2026-08-07 10:00:00+07"},
                {"id": 2, "source_path": "168_2024_nd_cp.C2.D6.K4", "target_path": "168_2024_nd_cp.C2.D6.K3.a", "ref_type": "dan_chieu", "created_at": "2026-08-07 10:01:00+07"},
                {"id": 3, "source_path": "168_2024_nd_cp.C1.D1.K2", "target_path": "168_2024_nd_cp.C1.D1.K1", "ref_type": "bo_sung", "created_at": "2026-08-07 10:02:00+07"},
                {"id": 4, "source_path": "168_2024_nd_cp.C10.D53.K1", "target_path": "45_2019_qh14.C11.D36.K1", "ref_type": "dan_chieu_cheo", "created_at": "2026-08-07 10:03:00+07"},
            ]
            fake_comments = [
                "Điểm b dẫn chiếu điểm a (sibling cùng Khoản)",
                "Khoản 4 dẫn chiếu Khoản 3 cùng Điều",
                "Khoản 2 bổ sung Khoản 1 (cùng Điều)",
                "Dẫn chiếu chéo: NĐ 168 → BLLĐ 2019 (khác doc)",
            ]
            for fi, fr in enumerate(fake_refs):
                rcols = ["id", "source_path", "target_path", "ref_type", "created_at"]
                for ci, col in enumerate(rcols, 1):
                    ws.cell(row, ci, _fmt(fr[col])).font = NORMAL
                    ws.cell(row, ci).border = BORDER
                    ws.cell(row, ci).alignment = WRAP
                ws.cell(row, COMMENT_COL, fake_comments[fi]).font = Font(size=9)
                ws.cell(row, COMMENT_COL).fill = YELLOW
                ws.cell(row, COMMENT_COL).alignment = WRAP
                ws.cell(row, COMMENT_COL).border = BORDER
                row += 1

        row += 1  # blank row between tables

    # ── column widths ──
    for c in range(1, COMMENT_COL):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions[get_column_letter(COMMENT_COL)].width = 42

    # ── freeze panes ──
    ws.freeze_panes = "B4"

    cur.close()
    conn.close()
    return wb


if __name__ == "__main__":
    wb = build()
    out = "docs/Sample_DB_v2.xlsx"
    try:
        wb.save(out)
        print(f"[OK] Saved: {out}")
    except PermissionError:
        out2 = "docs/Sample_DB_v2_new.xlsx"
        wb.save(out2)
        print(f"[OK] Saved: {out2}  (original locked by Excel)")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Rows: {wb.active.max_row}, Cols: {wb.active.max_column}")
